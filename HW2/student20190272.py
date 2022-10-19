#!/usr/bin/python3

import openpyxl
wb =openpyxl.load_workbook('student.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')


for i in range (len(sheet['a'])-1):
  midterm_score = sheet.cell(row=2+i, column=3).value
  final_score = sheet.cell(row=2+i, column=4).value
  homework_score = sheet.cell(row=2+i, column=5).value
  attendance_score = sheet.cell(row=2+i, column=6).value


  sheet.cell(row=2+i, column=7).value = midterm_score * 0.3 + final_score * 0.35 + homework_score * 0.34 + attendance_score * 1

  total_score = sheet.cell(row=2+i, column=7).value

 
A_plus = int(0.15 * (len(sheet['a'])-1))
A_zero = int(0.30 * (len(sheet['a'])-1))
B_plus = int(0.40 * (len(sheet['a'])-1))
B_zero = int(0.50 * (len(sheet['a'])-1))
C_plus = int(0.75 * (len(sheet['a'])-1))
C_zero = int(1.00 * (len(sheet['a'])-1))

 
score = []
result = []
for i in range (len(sheet['a'])-1):
  score.append(sheet.cell(row=2+i, column=7).value)
 

for i in range(0, len(score)):
    r = 1
    for j in range(0, len(score)):
        if score[i] < score[j]: r += 1
    result.append(r)

 

for i in range(len(result)):
  if result[i] <= A_plus:
    sheet.cell(row=2+i, column=8).value = "A+"
  elif result[i] > A_plus and result[i] <= A_zero:
    sheet.cell(row=2+i, column=8).value = "A0"
  elif result[i] > A_zero and result[i] <= B_plus:
    sheet.cell(row=2+i, column=8).value = "B+"
  elif result[i] > B_plus and result[i] <= B_zero:
    sheet.cell(row=2+i, column=8).value = "B0"
  elif result[i] > B_zero and result[i] <= C_plus:
    sheet.cell(row=2+i, column=8).value = "C+"
  elif result[i] > C_plus and result[i] <= C_zero:
    sheet.cell(row=2+i, column=8).value = "C0"

 
wb.save('student.xlsx')



